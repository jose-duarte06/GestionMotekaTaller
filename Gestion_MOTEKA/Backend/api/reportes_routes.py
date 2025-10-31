from flask import Blueprint, request, send_file, jsonify
from flask_jwt_extended import jwt_required
from core.auth import role_required
from models.ordenes import OrdenTrabajo, EstadoOrdenEnum
from models.vehiculos import Motocicleta
from models.personas import Cliente, Empleado
from models.catalogos import ModeloMoto, MarcaMoto
from datetime import datetime
import csv
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

reportes_bp = Blueprint('reportes', __name__, url_prefix='/api/reportes')

def get_ordenes_filtered(request_args):
    """Obtiene órdenes con los mismos filtros que el endpoint de órdenes"""
    cliente_id = request_args.get('cliente_id', type=int)
    cliente_nombre = request_args.get('cliente_nombre', '').strip()
    motocicleta_id = request_args.get('motocicleta_id', type=int)
    mecanico_id = request_args.get('mecanico_id', type=int)
    estado = request_args.get('estado', '').strip()
    desde = request_args.get('desde', '').strip()
    hasta = request_args.get('hasta', '').strip()
    placa = request_args.get('placa', '').strip()
    
    query = OrdenTrabajo.query.join(Cliente).join(Motocicleta).outerjoin(Empleado)
    
    if cliente_id:
        query = query.filter(OrdenTrabajo.cliente_id == cliente_id)
    
    if cliente_nombre:
        query = query.filter(Cliente.nombre.ilike(f'%{cliente_nombre}%'))
    
    if motocicleta_id:
        query = query.filter(OrdenTrabajo.motocicleta_id == motocicleta_id)
    
    if mecanico_id:
        query = query.filter(OrdenTrabajo.mecanico_asignado_id == mecanico_id)
    
    if estado:
        try:
            estado_enum = EstadoOrdenEnum[estado]
            query = query.filter(OrdenTrabajo.estado == estado_enum)
        except KeyError:
            pass
    
    if desde:
        try:
            fecha_desde = datetime.fromisoformat(desde.replace('Z', '+00:00'))
            query = query.filter(OrdenTrabajo.fecha_ingreso >= fecha_desde)
        except:
            pass
    
    if hasta:
        try:
            fecha_hasta = datetime.fromisoformat(hasta.replace('Z', '+00:00'))
            query = query.filter(OrdenTrabajo.fecha_ingreso <= fecha_hasta)
        except:
            pass
    
    if placa:
        query = query.filter(Motocicleta.placa.ilike(f'%{placa}%'))
    
    return query.order_by(OrdenTrabajo.fecha_ingreso.desc()).all()


@reportes_bp.route('/ordenes', methods=['GET'])
@jwt_required()
@role_required('gerente', 'encargado')
def exportar_ordenes():
    formato = request.args.get('formato', 'csv').lower()
    
    ordenes = get_ordenes_filtered(request.args)
    
    if formato == 'csv':
        return generar_csv(ordenes)
    elif formato == 'xlsx':
        return generar_xlsx(ordenes)
    elif formato == 'pdf':
        try:
            return generar_pdf(ordenes)
        except Exception as e:
            return generar_csv(ordenes)
    else:
        return jsonify({"error": "Formato no soportado. Use: csv, xlsx o pdf"}), 400


def generar_csv(ordenes):
    """Genera CSV con UTF-8 BOM"""
    output = io.StringIO()
    output.write('\ufeff')
    
    writer = csv.writer(output)
    writer.writerow(['ID', 'Cliente', 'Motocicleta (Placa)', 'Marca', 'Modelo', 'Estado', 
                    'Mecánico', 'Fecha Ingreso', 'Fecha Salida', 'Observaciones'])
    
    for orden in ordenes:
        moto = orden.motocicleta
        marca = moto.modelo.marca.nombre if moto and moto.modelo and moto.modelo.marca else 'N/A'
        modelo = moto.modelo.nombre if moto and moto.modelo else 'N/A'
        placa = moto.placa if moto else 'N/A'
        mecanico = orden.mecanico_asignado.nombre if orden.mecanico_asignado else 'Sin asignar'
        
        writer.writerow([
            orden.id,
            orden.cliente.nombre if orden.cliente else '',
            placa,
            marca,
            modelo,
            orden.estado.value if orden.estado else '',
            mecanico,
            orden.fecha_ingreso.strftime('%Y-%m-%d %H:%M') if orden.fecha_ingreso else '',
            orden.fecha_salida.strftime('%Y-%m-%d %H:%M') if orden.fecha_salida else '',
            orden.observaciones or ''
        ])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='ordenes_trabajo.csv'
    )


def generar_xlsx(ordenes):
    """Genera archivo XLSX con openpyxl"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes de Trabajo"
    
    headers = ['ID', 'Cliente', 'Motocicleta (Placa)', 'Marca', 'Modelo', 'Estado', 
               'Mecánico', 'Fecha Ingreso', 'Fecha Salida', 'Observaciones']
    ws.append(headers)
    
    for orden in ordenes:
        moto = orden.motocicleta
        marca = moto.modelo.marca.nombre if moto and moto.modelo and moto.modelo.marca else 'N/A'
        modelo = moto.modelo.nombre if moto and moto.modelo else 'N/A'
        placa = moto.placa if moto else 'N/A'
        mecanico = orden.mecanico_asignado.nombre if orden.mecanico_asignado else 'Sin asignar'
        
        ws.append([
            orden.id,
            orden.cliente.nombre if orden.cliente else '',
            placa,
            marca,
            modelo,
            orden.estado.value if orden.estado else '',
            mecanico,
            orden.fecha_ingreso.strftime('%Y-%m-%d %H:%M') if orden.fecha_ingreso else '',
            orden.fecha_salida.strftime('%Y-%m-%d %H:%M') if orden.fecha_salida else '',
            orden.observaciones or ''
        ])
    
    for column in range(1, len(headers) + 1):
        col_letter = get_column_letter(column)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='ordenes_trabajo.xlsx'
    )


def generar_pdf(ordenes):
    """Genera PDF horizontal con reportlab"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    elements = []
    
    data = [['ID', 'Cliente', 'Placa', 'Marca', 'Modelo', 'Estado', 'Mecánico', 'Ingreso', 'Salida']]
    
    for orden in ordenes:
        moto = orden.motocicleta
        marca = moto.modelo.marca.nombre if moto and moto.modelo and moto.modelo.marca else 'N/A'
        modelo = moto.modelo.nombre if moto and moto.modelo else 'N/A'
        placa = moto.placa if moto else 'N/A'
        mecanico = orden.mecanico_asignado.nombre if orden.mecanico_asignado else 'Sin asignar'
        
        data.append([
            str(orden.id),
            orden.cliente.nombre[:20] if orden.cliente else '',
            placa[:15],
            marca[:15],
            modelo[:15],
            orden.estado.value if orden.estado else '',
            mecanico[:20],
            orden.fecha_ingreso.strftime('%Y-%m-%d') if orden.fecha_ingreso else '',
            orden.fecha_salida.strftime('%Y-%m-%d') if orden.fecha_salida else ''
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    output.seek(0)
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name='ordenes_trabajo.pdf'
    )
